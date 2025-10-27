import pandas as pd
from pathlib import Path
from bs4 import BeautifulSoup
from logger_setup import setup_logger
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
if __package__ is None or __package__ == "":
    import sys, pathlib
    sys.path.append(str(pathlib.Path(__file__).resolve().parent))


setup_logger("parse_page.log")

data_dir = Path(__file__).parent / 'data'
html_path = data_dir / 'page0.html'
BASE_DIR = Path(__file__).parent.resolve()

def safe_text(el):
    """Безопасно берёт текст из тега.
    Возвращает '' если элемента нет, и сжимает пробелы."""
    if not el:
        return ""
    return " ".join(el.get_text(" ", strip=True).split())


try:
    html = html_path.read_text(encoding='utf-8')
except FileNotFoundError:
    print("Файл не найден. Сначала требуется запусить fetch_page.py")
    exit(1)
try:
    soup = BeautifulSoup(html, 'lxml')

    vacancies = []

    for title_el in soup.select('a[data-qa="serp-item__title"]'):
        card = title_el.find_parent(
            lambda t: t and t.has_attr('data-qa') and 'vacancy-serp' in t['data-qa']
        )
        if not card:
            continue
        company_el = (
            card.select_one('[data-qa="vacancy-serp__vacancy-employer"]')
            if card else title_el.find_next(attrs={'data-qa': 'vacancy-serp__vacancy-employer'})
        )
        exp_el = card.select_one('[data-qa*="vacancy-work-experience"]')
        vacancies.append({
            'Название вакансии': safe_text(title_el),
            'Компания': safe_text(company_el),
            'Опыт': safe_text(exp_el),  # ← новое поле
        })
        if len(vacancies) >= 30:
            break

except Exception as e:
    print("Ошибка парсинга", e)

df = pd.DataFrame(vacancies)
df = df.applymap(lambda x: " ".join(x.split()) if isinstance(x, str) else x)
df = df.replace({"": "-"}).fillna("-")
cols = ["Название вакансии", "Компания", "Опыт"]
df = df.reindex(columns=cols)
out_csv = BASE_DIR / 'data' / 'hh_kazan_qa.csv'
df.to_csv(out_csv, index=False, encoding='utf-8-sig')
print('CSV:', out_csv, '| rows =', len(df))

xlsx_path = BASE_DIR / 'data' / 'hh_kazan_qa.xlsx'

with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Vacancies')
    ws = writer.sheets['Vacancies']

    # 1) стилизуем заголовки
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='solid', start_color='FFEFEFEF', end_color='FFEFEFEF')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # 2) автоширина колонок (с ограничениями по разумной ширине)
    for idx, col in enumerate(df.columns, start=1):
        max_len = max(len(str(col)), *(len(str(v)) for v in df[col].astype(str)))
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(max_len + 2, 60))

    # 3) мелочи для удобства
    ws.freeze_panes = 'A2'              # фиксируем шапку
    ws.auto_filter.ref = ws.dimensions  # автофильтр по всей таблице

print('XLSX:', xlsx_path, '| rows =', len(df))